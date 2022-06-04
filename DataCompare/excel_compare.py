import pandas as pd
import os
import pathlib
import datetime as dt
from openpyxl import Workbook, load_workbook
import function_def as func
from flask import Flask , request
from flask_cors import CORS, cross_origin

def compare(result_file_name):
    print("Init result file ++++++++++++++++++++++++++++++++++++", result_file_name)
    config_sheet = result_file_name
    flag = False
    result_file_name = "Result.xlsx"
    result_wb = Workbook()
    result_wb.remove(result_wb["Sheet"])
    index_sheet = init_index_sheet(result_wb)
    parent_dir = os.path.dirname(os.path.abspath(__file__))
    for row_index in range(2, config_sheet.max_row + 1):
        print(f'Run: {row_index - 1}')
        file1_name = config_sheet["A{}".format(row_index)].value
        file2_name = config_sheet["B{}".format(row_index)].value
        file1_path = f'{parent_dir}//golden_output//{file1_name}.xlsx'
        file2_path = f'{parent_dir}//output//{file2_name}.xlsx'
        file1_sheet = load_workbook(file1_path).worksheets[0]
        file2_sheet = load_workbook(file2_path).worksheets[0]
        start_time = dt.datetime.now()

        print("comparing", file1_name, "and", file2_name)
        print("start time: ", start_time)
        keys = config_sheet["C{}".format(row_index)].value
        # keys = get_selected_keys(keys_header)
        col_to_compare = config_sheet["D{}".format(row_index)].value
        col_to_ignore_file1 = config_sheet["E{}".format(row_index)].value
        col_to_ignore_file2 = config_sheet["F{}".format(row_index)].value

        if keys in "No Key Provided":
            index_sheet.cell(row=row_index, column=1, value=row_index - 1)
            index_sheet.cell(row=row_index, column=2, value=file1_name)
            index_sheet.cell(row=row_index, column=3, value=file2_name)
            index_sheet.cell(row=row_index, column=6, value="Warning No Key Provided")
            func.format_sheet_row_warning(index_sheet, row_index, 6)
            end_time = dt.datetime.now()
            print("end time ", end_time)
            print("compare time: ", end_time - start_time)
        else:
            keys_count = len(keys.split(","))
            file1_df = get_dataframe(file1_path, keys, file1_sheet, "default")
            file2_df = get_dataframe(file2_path, keys, file2_sheet, "default")
            if ("Invalid key" in file1_df) or ("Invalid key" in file2_df):
                index_sheet.cell(row=row_index, column=1, value=row_index - 1)
                index_sheet.cell(row=row_index, column=2, value=file1_name)
                index_sheet.cell(row=row_index, column=3, value=file2_name)
                index_sheet.cell(row=row_index, column=6, value=f'Warning {file1_df}')
                func.format_sheet_row_warning(index_sheet, row_index, 6)
                end_time = dt.datetime.now()
                print("end time ", end_time)
                print("compare time: ", end_time - start_time)
            else:
                all_cols = get_columns_union(file1_df, file2_df, keys)

                common_indexes = []
                deleted_indexes = []
                new_indexes = []

                for index in file1_df.index:
                    if index in file2_df.index:
                        common_indexes.append(index)
                    else:
                        deleted_indexes.append(index)

                for index in file2_df.index:
                    if index not in file1_df.index:
                        new_indexes.append(index)

                prepare_summary_sheet(row_index-1,result_wb, file1_name, file2_name,file1_sheet,file2_sheet,common_indexes,deleted_indexes,new_indexes)
                details_sheet = init_detail_sheet(row_index-1,result_wb,keys.split(","),all_cols)
                any_diff = compare_common_indexes(common_indexes,details_sheet,file1_df,file2_df,all_cols,col_to_ignore_file1,col_to_ignore_file2)

                populate_indexes(new_indexes,details_sheet,file2_df,all_cols,False)
                result_wb.save(result_file_name)
                populate_indexes(deleted_indexes,details_sheet,file1_df,all_cols,True)

                index_sheet.cell(row=row_index, column=1, value=row_index - 1)
                index_sheet.cell(row=row_index, column=2, value=file1_name)
                index_sheet.cell(row=row_index, column=3, value=file2_name)
                index_sheet.cell(row=row_index, column=4, value=f'=HYPERLINK("#SummarySheet{row_index-1}!A1","Summary")')
                index_sheet.cell(row=row_index, column=5,value=f'=HYPERLINK("#DetailSheet{row_index - 1}!A1","Details")')
                if flag is True:
                    func.fill_index_dataCellDark(index_sheet, row_index, 1)
                    func.fill_index_dataCellDark(index_sheet, row_index, 2)
                    func.fill_index_dataCellDark(index_sheet, row_index, 3)
                    func.fill_index_dataCellDark(index_sheet, row_index, 4)
                    func.fill_index_dataCellDark(index_sheet, row_index, 5)
                    func.fill_index_dataCellDark(index_sheet, row_index, 6)
                    flag = False
                else:
                    func.fill_index_dataCellLight(index_sheet, row_index, 1)
                    func.fill_index_dataCellLight(index_sheet, row_index, 2)
                    func.fill_index_dataCellLight(index_sheet, row_index, 3)
                    func.fill_index_dataCellLight(index_sheet, row_index, 4)
                    func.fill_index_dataCellLight(index_sheet, row_index, 5)
                    func.fill_index_dataCellLight(index_sheet, row_index, 6)
                    flag = True



                if any_diff or len(new_indexes) > 0 or len(deleted_indexes) > 0:
                    index_sheet.cell(row=row_index,column=6,value="FAIL")
                    func.format_cell_status_fail(index_sheet,row_index,6)
                    func.format_border_thin(index_sheet,row_index,6)
                else:
                    index_sheet.cell(row=row_index, column=6, value="PASS")
                    func.format_cell_status_pass(index_sheet, row_index, 6)
                    func.format_border_thin(index_sheet, row_index, 6)

                details_sheet.auto_filter.ref = 'B2'
                index_sheet.auto_filter.ref = 'F1'

                end_time = dt.datetime.now()
                print("end time :", end_time)
                print(("compare time :",end_time-start_time))

    result_wb.save(result_file_name)
    result_wb.save("ResultWithIndex.xlsx")
    print("Saving Result file")

    deleteAllSheetsExcept("ResultWithIndex.xlsx")
    return result_file_name
def compare_common_indexes(common_indexes, details_sheet,file1_df,file2_df,all_cols,cols_to_ignore_file1,cols_to_ignore_file2):

    col_index = {}
    for col in all_cols:
        col_index[str(col)] = get_column_index(col,details_sheet)

    cols_to_ignore = []
    if cols_to_ignore_file1 != None:
        cols_to_ignore = cols_to_ignore + cols_to_ignore_file1.split(",")
    if cols_to_ignore_file2 != None:
        cols_to_ignore = cols_to_ignore + cols_to_ignore_file2.split(",")

    any_diff = False
    for index in common_indexes:
        row_idx = populate_index(index,details_sheet,None)
        row1 = file1_df.loc[index]
        row2 = file2_df.loc[index]
        status = "Matched"
        for col in all_cols:
            val1 = row1[col]
            val2 = row2[col]
            details_sheet.cell(row=row_idx,column=col_index[col],value=val1)
            details_sheet.cell(row=row_idx,column=col_index[col]+1, value=val2)

            if format(val1) != format(val2) and col not in cols_to_ignore:
                status = "Not Matched"
                any_diff = True
                func.format_sheet_cell_un_matched(details_sheet,row_idx,col_index[col])
                func.format_sheet_cell_un_matched(details_sheet,row_idx,col_index[col]+1)

        details_sheet.cell(row=row_idx,column=2,value=status)
        if status != "Matched":
            func.format_sheet_cell_un_matched(details_sheet, row_idx,2)

    return any_diff

def format(val):
    if val!=val:
        return ""
    else:
        return val

def init_index_sheet(result_wb):
    print("Init Index Sheet")
    index_sheet = result_wb.create_sheet("Index")
    index_sheet.cell(row=1,column=1,value="Index")
    func.format_index_headers(index_sheet,1,1)
    index_sheet.cell(row=1,column=2,value="File 1")
    func.format_index_headers(index_sheet,1,2)
    index_sheet.cell(row=1,column=3,value="File 2")
    func.format_index_headers(index_sheet,1,3)
    index_sheet.cell(row=1,column=4,value="Summary Sheet")
    func.format_index_headers(index_sheet,1,4)
    index_sheet.cell(row=1,column=5,value="Detail Sheet")
    func.format_index_headers(index_sheet,1,5)
    index_sheet.cell(row=1,column=6,value="Status")
    func.format_index_headers(index_sheet,1,6)

    return index_sheet

def formatsheetheaders(wb,ws):
    func.format_index_headers(ws,1,1)
    func.format_index_headers(ws,1,2)
    func.format_index_headers(ws,1,3)
    func.format_index_headers(ws,1,4)
    func.format_index_headers(ws,1,5)
    func.format_index_headers(ws,1,6)

def prepare_summary_sheet(sheet_number,result_wb,file1_name,file2_name,file1_sheet,file2_sheet,common_indexes,deleted_indexes,new_indexes):
    summary_sheet = result_wb.create_sheet("SummarySheet{}".format(sheet_number))
    summary_sheet.cell(row=1,column=1,value="")
    func.format_sheet_headers(summary_sheet,1,1)
    summary_sheet.cell(row=1,column=2,value="File 1")
    func.format_sheet_headers(summary_sheet, 1, 2)
    summary_sheet.cell(row=1, column=3, value="File 2")
    func.format_sheet_headers(summary_sheet, 1, 3)

    summary_sheet.cell(row=2, column=1, value="Name")
    func.format_sheet_SummaryValue(summary_sheet, 2, 1)
    func.format_sheet_SummaryValue(summary_sheet, 2, 2)
    func.format_sheet_SummaryValue(summary_sheet, 2, 3)
    summary_sheet.cell(row=2, column=2, value=file1_name)
    summary_sheet.cell(row=2, column=3, value=file2_name)

    summary_sheet.cell(row=3, column=1, value="Row Count")
    func.format_sheet_SummaryValue(summary_sheet, 3, 1)
    func.format_sheet_SummaryValue(summary_sheet, 3, 2)
    func.format_sheet_SummaryValue(summary_sheet, 3, 3)
    summary_sheet.cell(row=3, column=2, value=file1_sheet.max_row-1)
    summary_sheet.cell(row=3, column=3, value=file2_sheet.max_row-1)

    summary_sheet.cell(row=4, column=1, value="Column Count")
    func.format_sheet_SummaryValue(summary_sheet, 4, 1)
    func.format_sheet_SummaryValue(summary_sheet, 4, 2)
    func.format_sheet_SummaryValue(summary_sheet, 4, 3)
    summary_sheet.cell(row=4, column=2, value=file1_sheet.max_column - 1)
    summary_sheet.cell(row=4, column=3, value=file2_sheet.max_column - 1)

    summary_sheet.cell(row=5, column=1, value="Matching Key Rows")
    func.format_sheet_SummaryValue(summary_sheet, 5, 1)
    func.format_sheet_SummaryValue(summary_sheet, 5, 2)
    summary_sheet.cell(row=5, column=2, value=len(common_indexes))
    summary_sheet.merge_cells(start_row=5, start_column=2, end_row=5,end_column=3)

    summary_sheet.cell(row=6, column=1, value="Deleted Rows")
    func.format_sheet_SummaryValue(summary_sheet, 6, 1)
    func.format_sheet_SummaryValue(summary_sheet, 6, 2)
    summary_sheet.cell(row=6, column=2, value=len(deleted_indexes))
    summary_sheet.merge_cells(start_row=6, start_column=2, end_row=6,end_column=3)

    summary_sheet.cell(row=7, column=1, value="New Rows")
    func.format_sheet_SummaryValue(summary_sheet, 7, 1)
    func.format_sheet_SummaryValue(summary_sheet, 7, 2)
    summary_sheet.cell(row=7, column=2, value=len(new_indexes))
    summary_sheet.merge_cells(start_row=7, start_column=2, end_row=7,end_column=3)

    summary_sheet.cell(row=10, column=1, value=f'=HYPERLINK("#Index!A1","Back To Index")').hyperlink
    summary_sheet.merge_cells(start_row=10, start_column=1, end_row=10, end_column=3)
    func.format_hyperlink(summary_sheet,10,1)

    return summary_sheet

def init_detail_sheet(sheet_number,result_wb,keys,all_cols):
    #print("init details sheet",sheet_number)
    details_sheet = result_wb.create_sheet("DetailSheet{}".format(sheet_number))

    details_sheet.cell(row=1,column=1,value="S. No")
    func.format_sheet_headers(details_sheet,1,1)
    details_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    details_sheet.cell(row=1, column=2, value="Status Type")
    func.format_sheet_headers(details_sheet, 1, 2)
    details_sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    col_index = 2
    for key in keys:
        col_index = col_index+1
        details_sheet.cell(row=1, column=col_index, value=key)
        func.format_sheet_headers(details_sheet, 1, col_index)
        details_sheet.merge_cells(start_row=1, start_column=col_index, end_row=2, end_column=col_index)

    cod = details_sheet.cell(3,col_index+1).coordinate

    for col in all_cols:
        col_index = col_index+1
        details_sheet.cell(row=1,column=col_index,value=col)
        func.format_sheet_headers(details_sheet,1,col_index)
        details_sheet.cell(row=2, column=col_index,value="Old")
        func.format_sheet_headers(details_sheet,2,col_index)
        details_sheet.cell(row=2, column=col_index+1,value="New")
        func.format_sheet_headers(details_sheet, 2, col_index+1)
        details_sheet.merge_cells(start_row=1,start_column=col_index,end_row=1,end_column=col_index+1)
        col_index = col_index+1

    details_sheet.freeze_panes = cod
    return details_sheet

def get_columns_union(file1_df,file2_df,keys):
    detail_cols = []
    for col in file1_df.columns:
        if col not in detail_cols and col not in keys:
            detail_cols.append(col)
    for col in file2_df.columns:
        if col not in detail_cols and col not in keys:
            detail_cols.append(col)
    return detail_cols

def get_dataframe(file_path, keys, sheet, sheetname):
    index_list = []
    for key in keys.split(","):
        index = get_column_index(key, sheet)
        if index is None:
            contents = "Invalid key: "+key
            return contents
        else:
            index_list.append(index - 1) #index is 0 based

    if sheetname == "default":
        contents = pd.read_excel(file_path)
        #print(contents)
    else:
        contents = pd.read_excel(file_path,sheet_name=sheetname)
    contents.fillna('',inplace=True)
    contents.set_index(keys.split(","), inplace=True)
    return contents

def get_column_index(col_name,sheet):
    for col_idx in range(1, sheet.max_column+1):
        cell = sheet.cell(row=1,column=col_idx)
        if(cell.value == col_name) or (cell.value is not None and col_name is not None and cell.value.lower() == col_name.lower()):
            return col_idx

def populate_index(index, details_sheet, status):
    row_index = details_sheet.max_row +1
    details_sheet.cell(row=row_index,column=1, value=row_index - 3)
    if status is not None:
        details_sheet.cell(row=row_index,column=2, value=status)
        if status == "New":
            func.format_sheet_row_new(details_sheet, row_index, 2)
        else:
            func.format_sheet_row_removed(details_sheet,row_index, 2)

    if type(index) is tuple:
        for col_index, kvalue in enumerate(index, start=3):
            #details_sheet.cell(row=row_index,column=col_index, value=index)
            details_sheet.cell(row=row_index,column=col_index,value=kvalue)
            #func.format_border_thin(details_sheet, row_index, col_index)
    else:
        for col_index, kvalue in enumerate(index, start=3):
            details_sheet.cell(row=row_index,column=col_index,value=index)
            #func.format_border_thin(details_sheet, row_index, col_index)
            #result_wb.save(result_file_name)

    return row_index

def populate_index_NewRemoved(index, details_sheet, status):
    row_index = details_sheet.max_row +1
    details_sheet.cell(row=row_index,column=1, value=row_index - 3)
    if status is not None:
        details_sheet.cell(row=row_index,column=2, value=status)
        if status == "New":
            func.format_sheet_row_new(details_sheet, row_index, 2)
        else:
            func.format_sheet_row_removed(details_sheet,row_index, 2)

    if type(index) is tuple:
        for col_index, kvalue in enumerate(index, start=3):
            details_sheet.cell(row=row_index, column=col_index, value=kvalue)
            break

    else:
        for col_index, kvalue in enumerate(index, start=3):
            details_sheet.cell(row=row_index, column=col_index, value=index)
            break


    return row_index

def populate_indexes(indexes, details_sheet, df, all_cols, removed):
    cols_index = {}
    for col in all_cols:
        cols_index[str(col)] = get_column_index(col, details_sheet)

    status = "New"
    if removed:
        status = "Removed"

    for index in indexes:
        row_idx = populate_index_NewRemoved(index, details_sheet, status)
        row = df.loc[index]
        for col in all_cols:
            if removed:
                details_sheet.cell(row=row_idx,column=cols_index[(col)], value=row[(col)])
            else:
                details_sheet.cell(row=row_idx, column=cols_index[(col)] + 1, value=row[(col)])


def deleteAllSheetsExcept(xlname):
    wb = load_workbook(xlname)
    sheet_list = wb.sheetnames
    for name in sheet_list:
        if name not in "Index":
            wb.remove(wb[name])

    wb.save(xlname)

#def get_selected_keys(headers,sheetname):



if __name__ == "__main__":
    master_file_name = "Master.xlsx"
    start_time = dt.datetime.now()
    print("Excel Comparsion Start at: ",start_time)
    for master_file in pathlib.Path.cwd().glob(master_file_name):
        master_file_name = master_file_name

    master_wb = load_workbook(os.path.join(os.getcwd(),master_file))
    config_sheet = master_wb["config"]
    compare(config_sheet)
    end_time = dt.datetime.now
    print("Excel Comparsion End at: ", end_time)
    print("Total Comparsion time: ", str(dt.datetime.now() - start_time))
    print("Done !")





































