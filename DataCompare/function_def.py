import glob
import json
import os
import pathlib
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import colors, Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, BORDER_THICK, BORDER_THIN, Side
from openpyxl.styles.fills import PatternFill


def create_excel(name):
    wb = Workbook()
    wb.save(name)
    return wb

def load_excel(name):
    wb = load_workbook(name)
    return wb

def get_current_working_dir():
    path = pathlib.Path.cwd()
    return path

def print_sheet(filename, sheetIndex):
    wb= load_workbook(filename + ".xlsx")
    ws= wb.worksheets[sheetIndex]
    for row in range(1, get_maximum_row(ws)):
        for column in range(1, get_maximum_column(ws)):
            item = ws.cell(row=row,column=column)
            print(item.value,end=" ")
    print("*" * 100)

def get_maximum_row(ws):
    row = ws.max_row
    return row

def get_maximum_column(ws):
    column = ws.max_column
    return column

def read_value_by_cell(wb, sheet, cell):
    ws = wb.worksheets[sheet]
    return ws[cell].value

def wbread_value_from_cell(wb, sheet, cell):
    ws = wb[sheet]
    return ws[cell].value


def search_column(ws, searchedItem):
    counter = 1
    flag = False
    for column in range(1, get_maximum_column(ws) + 1):
        item = ws.cell(row=1, column=column)
        if (item.value == searchedItem) or (item.value is not None and searchedItem is not None and item.value.lower() == searchedItem.lower()):
            #print("Column {} is found at {}".format(searchedItem,counter))
            flag = True
            break
        else:
            counter + 1

    if(flag==False):
        column = get_maximum_column(ws) + 1
        return column

    return counter

def search_row(ws,col,searchedItem):
    counter = 1
    flag = False
    for row in range(1, get_maximum_row(ws) + 1):
        item = ws.cell(row=row,column=col)

        if(item.value == searchedItem):
            #print("Column {} is found at {}".format(searchedItem,counter))
            flag=True
            break
        else:
            counter +=1
    if(flag==False):
        column = get_maximum_column(ws) + 1
        return column

    return counter
    print("*" * 100)



def list_columns(ws, keyword):
    list_c = []
    counter = 1
    flag = False
    for column in range(1, get_maximum_column(ws) + 1):
        item = ws.cell(row=1, column=column)

        if (keyword in item.value):
            list_c.append(item.value)

    return list_c


def format_cell_header(ws, rowNum, colNum):
    ws.cell(row=rowNum, column=colNum).font = Font(name="Calibri", b=True, i=True, size=12)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center",vertical=None, textRotation=0,wrapText=None,\
                                                             indent=0,relativeIndent=0,\
                                                             justifyLastLine=None,readingOrder=0,text_rotation=None,\
                                                             wrap_text=None,shrink_to_fit=None,\
                                                             mergeCell=None)

def format_cell_column(ws, rowNum, colNum):
    ws.cell(row=rowNum, column=colNum).font = Font(name="Calibri", b=True, size=11)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center",vertical=None, textRotation=0,wrapText=None,\
                                                             indent=0,relativeIndent=0,\
                                                             justifyLastLine=None,readingOrder=0,text_rotation=None,\
                                                             wrap_text=None,shrink_to_fit=None,\
                                                             mergeCell=None)


def format_cell_column_fail(ws, rowNum, colNum):
    ws.cell(row=rowNum, column=colNum).font = Font(name="Calibri", b=True, size=11)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center",vertical=None, textRotation=0,wrapText=None,\
                                                             indent=0,relativeIndent=0,\
                                                             justifyLastLine=None,readingOrder=0,text_rotation=None,\
                                                             wrap_text=None,shrink_to_fit=None,\
                                                             mergeCell=None)


def format_cell_status_fail(ws, rowNum, colNum):
    ws.cell(row=rowNum, column=colNum).font = Font(name="Calibri",color="ff0000" ,b=True, size=11)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center",vertical=None, textRotation=0,wrapText=None,\
                                                             indent=0,relativeIndent=0,\
                                                             justifyLastLine=None,readingOrder=0,text_rotation=None,\
                                                             wrap_text=None,shrink_to_fit=None,\
                                                             mergeCell=None)


def format_cell_status_pass(ws, rowNum, colNum):
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center", vertical=None, textRotation=0,
                                                             wrapText=None, \
                                                             indent=0, relativeIndent=0, \
                                                             justifyLastLine=None, readingOrder=0, text_rotation=None, \
                                                             wrap_text=None, shrink_to_fit=None, \
                                                             mergeCell=None)

def format_border_thick(ws, rowNum, colNum):
    borders = Border(left=Side(border_style=BORDER_THICK, color='00000000'),
                     right=Side(border_style=BORDER_THICK, color='00000000'),
                     top=Side(border_style=BORDER_THICK, color='00000000'),
                     bottom=Side(border_style=BORDER_THICK, color='00000000'))
    ws.cell(row=rowNum, column=colNum).border = borders

def format_border_thin(ws, rowNum, colNum):
    borders = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
                     right=Side(border_style=BORDER_THIN, color='00000000'),
                     top=Side(border_style=BORDER_THIN, color='00000000'),
                     bottom=Side(border_style=BORDER_THIN, color='00000000'))
    ws.cell(row=rowNum, column=colNum).border = borders


def fill_red_colour(ws, rowNum, colNum):
    red_color = colors.Color(rgb='FF6666')
    solid_red_fill = PatternFill(patternType='solid', fgColor=red_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_red_fill

def fill_blue_colour(ws, rowNum, colNum):
    blue_color = colors.Color(rgb='6699cc')
    solid_blue_fill = PatternFill(patternType='solid', fgColor=blue_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_blue_fill

def fill_yellow_colour(ws, rowNum, colNum):
    yellow_color = colors.Color(rgb='ffff99')
    yellow_red_fill = PatternFill(patternType='solid', fgColor=yellow_color)
    ws.cell(row=rowNum, column=colNum).fill = yellow_red_fill

def fill_orange_colour(ws, rowNum, colNum):
    orange_color = colors.Color(rgb='f3af73')
    orange_red_fill = PatternFill(patternType='solid', fgColor=orange_color)
    ws.cell(row=rowNum, column=colNum).fill = orange_red_fill

def fill_grey_colour(ws, rowNum, colNum):
    grey_color = colors.Color(rgb='dee2e6')
    solid_grey_fill = PatternFill(patternType='solid', fgColor=grey_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_grey_fill

def fill_light_grey_colour(ws, rowNum, colNum):
    light_grey_color = colors.Color(rgb='fbfcfc')
    solid_lightgrey_fill = PatternFill(patternType='solid', fgColor=light_grey_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_lightgrey_fill

def fill_green_colour(ws, rowNum, colNum):
    green_color = colors.Color(rgb='008000')
    solid_green_fill = PatternFill(patternType='solid', fgColor=green_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_green_fill

def fill_wood_colour(ws, rowNum, colNum):
    wood_color = colors.Color(rgb='96c9b7')
    solid_wood_fill = PatternFill(patternType='solid', fgColor=wood_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_wood_fill

def fill_light_blue_colour(ws, rowNum, colNum):
    light_blue_color = colors.Color(rgb='d0e2e3')
    light_blue_color_fill = PatternFill(patternType='solid', fgColor=light_blue_color)
    ws.cell(row=rowNum, column=colNum).fill = light_blue_color_fill

def fill_light_blue_colourl(ws, rowNum, colNum):
    light_blue_colourl = colors.Color(rgb='f0f6f6')
    light_blue_colourl_fill = PatternFill(patternType='solid', fgColor=light_blue_colourl)
    ws.cell(row=rowNum, column=colNum).fill = light_blue_colourl_fill

def format_index_headers(ws, rowNum, colNum):
    col = colors.Color(rgb='436f70')
    solid_fill = PatternFill(patternType='solid', fgColor= col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill
    ws.cell(row=rowNum, column=colNum).font = Font("Courier New", color=colors.WHITE, b=True, size=11)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center", vertical=None, textRotation=0,
                                                             wrapText=None,\
                                                             indent=0, relativeIndent=0, \
                                                             justifyLastLine=None, readingOrder=0, text_rotation=None, \
                                                             wrap_text=None, shrink_to_fit=None, \
                                                             mergeCell=None)

def format_sheet_headers(ws, rowNum, colNum):
    col = colors.Color(rgb='177791')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill
    ws.cell(row=rowNum, column=colNum).font = Font("Calibri", color=colors.WHITE, b=True, size=11)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center", vertical=None, textRotation=0,
                                                             wrapText=None,
                                                             indent=0, relativeIndent=0, \
                                                             justifyLastLine=None, readingOrder=0, text_rotation=None, \
                                                             wrap_text=None, shrink_to_fit=None, \
                                                             mergeCell=None)
    format_border_thin(ws, rowNum, colNum)


def format_sheet_SummaryValue(ws, rowNum, colNum):
    col = colors.Color(rgb='1c93b3')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill
    ws.cell(row=rowNum, column=colNum).font = Font("Calibri", color=colors.WHITE, b=True, size=11)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center", vertical=None, textRotation=0,
                                                             wrapText=None,
                                                             indent=0, relativeIndent=0, \
                                                             justifyLastLine=None, readingOrder=0, text_rotation=None, \
                                                             wrap_text=None, shrink_to_fit=None, \
                                                             mergeCell=None)
    format_border_thin(ws, rowNum, colNum)




def fill_index_dataCellDark(ws, rowNum, colNum):
    red_color = colors.Color(rgb='bfd8d9')
    solid_red_fill = PatternFill(patternType='solid', fgColor=red_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_red_fill
    format_border_thin(ws, rowNum, colNum)


def fill_index_dataCellLight(ws, rowNum, colNum):
    red_color = colors.Color(rgb='eff5f6')
    solid_red_fill = PatternFill(patternType='solid', fgColor=red_color)
    ws.cell(row=rowNum, column=colNum).fill = solid_red_fill
    format_border_thin(ws, rowNum, colNum)

def format_hyperlink(ws, rowNum, colNum):
    col = colors.Color(rgb='d6f1f8')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill
    ws.cell(row=rowNum, column=colNum).font = Font("Calibri", color=colors.BLUE, b=True, underline='single', size=11)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center", vertical=None, textRotation=0,
                                                             wrapText=None,
                                                             indent=0, relativeIndent=0, \
                                                             justifyLastLine=None, readingOrder=0, text_rotation=None, \
                                                             wrap_text=None, shrink_to_fit=None, \
                                                             mergeCell=None)


def format_sheet_row_matched(ws, rowNum, colNum):
    col = colors.Color(rgb='e7f7fb')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill
    ws.cell(row=rowNum, column=colNum).font = Font("Calibri", b=True, size=10)
    ws.cell(row=rowNum, column=colNum).alignment = Alignment(horizontal="center", vertical=None, textRotation=0,
                                                             wrapText=None,
                                                             indent=0, relativeIndent=0, \
                                                             justifyLastLine=None, readingOrder=0, text_rotation=None, \
                                                             wrap_text=None, shrink_to_fit=None, \
                                                             mergeCell=None)


def format_sheet_row_removed(ws, rowNum, colNum):
    col = colors.Color(rgb='ff0000')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill

def format_sheet_row_warning(ws, rowNum, colNum):
    col = colors.Color(rgb='F5B041')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill

def format_sheet_row_new(ws, rowNum, colNum):
    col = colors.Color(rgb='90ee90')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill

def format_sheet_cell_matched(ws, rowNum, colNum):
    col = colors.Color(rgb='e7f7fb')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill

def format_sheet_cell_un_matched(ws, rowNum, colNum):
    col = colors.Color(rgb='ffff00')
    solid_fill = PatternFill(patternType='solid', fgColor=col)
    ws.cell(row=rowNum, column=colNum).fill = solid_fill

def addPieChart(ws,minC,minR,maxC,maxR,RowNum):
    series = Reference(ws, min_col=minC,min_row=minR,max_col=maxC,max_row=maxR)
    pie_chart = PieChart()
    labels = Reference(ws,min_col=1,min_row=4,max_col=1,max_row=6)
    #legends = Reference(ws,min_col=3,min_row=4,max_col=3,max_row=7)
    pie_chart.add_data(series, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.title = "Percentage Chart"
    #pie_chart.add_data(legends)
    ws.add_chart(pie_chart,"A{}".format(RowNum))

def j_son_to_dictionary(strng):
    a = json.load(strng)
    #for k,v in a.items():
        #print(f'Key: {k} and Value {v}')
    return a

def dictionary_to_string(dictionaryObject):
    strObj = ""
    for k,v in dictionaryObject.items():
        if strObj is not "":
            strObj += ","
        strObj += v
    return strObj

def search_in_dictionary(dictionaryKV, searchKey):
    matchedValue = ""
    for itemKey, itemValue in dictionaryKV.items():
        if itemKey == matchedValue:
            matchedValue = itemValue
    return matchedValue

def search_key_in_string(keystring, searchedKey):
    result = "False"
    all_items = keystring.split(",")
    for itemKey in all_items:
        if searchedKey in itemKey:
            result = True
    return result

def get_result_path(folder, filename):
    absFilePath = os.path.abspath(__file__)
    print(absFilePath)
    parentDir = os.path.dirname(os.path.abspath(__file__))
    resultpath = f'{parentDir}//output//newOutput//mem_results_standard.xlsx'
    print(resultpath)

def cln_dir(directory):
    checkDir = pathlib.Path(directory)
    if checkDir.exists():
        shutil.rmtree(directory)
        os.mkdir(directory)
    else:
        os.mkdir(directory)

def cln_file_data(path):
    master_workbook = load_workbook(path)
    master_sheet = master_workbook.worksheets[0]
    count = get_maximum_row(master_sheet)
    for i in range(2, count + 1):
        master_sheet.delete_rows(2)
    master_workbook.save(path)

def update_excel(master_path, master_workbook, master_sheet, scenarioName, keys, pattern, columnToCompare, columnToIgnoreOld, columnToIgnoreNew):
    for name in glob.glob(pattern):
        splitName = name.split("output\\")
        fileName = splitName[1].split(".xlsx")
        keyName = fileName[0].split(f'{scenarioName}_')
        r = get_maximum_row(master_sheet)+1
        master_sheet.cell(row=r,column=1,value=fileName[0])
        master_sheet.cell(row=r,column=2,value=fileName[0])
        key = ""
        if keys is None:
            master_sheet.cell(row=r,column=3, value="No key Provided")
        else:
            dataToLoad = j_son_to_dictionary(keys)
            key = search_in_dictionary(dataToLoad, keyName[1])
            master_sheet.cell(row=r, column=3, value=key)
        master_sheet.cell(row=r,column=4, value=columnToCompare)
        master_sheet.cell(row=r,column=5, value=columnToIgnoreOld)
        master_sheet.cell(row=r,column=6, value=columnToIgnoreNew)
    master_workbook.save(master_path)

def string_to_json(stringname, wbController, SheetName, row):
    keys = "{"
    if stringname is None:
        print("None type")
    else:
        ws = wbController[SheetName]
        all_items = stringname.split(",")
        for item in all_items:
            rowcount = get_maximum_row(ws)
            for row in range (2, rowcount):
                rowkey = wbread_value_from_cell(wbController,SheetName, "A{}".format(row))
                rowValue = wbread_value_from_cell(wbController, SheetName, "B{}".format(row))
                if item in rowkey:
                    if (keys is "{"):
                        keys = keys + '"'+rowkey+'":"'+rowValue+'"'
                    else:
                        keys = keys + "," + '"'+rowkey+'":"'+rowValue + '"'
        keys = keys + "}"
    return keys








































































































































































